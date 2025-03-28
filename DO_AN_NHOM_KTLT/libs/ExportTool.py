from datetime import datetime

from openpyxl import load_workbook
import xlsxwriter as xr

from DO_AN_NHOM_KTLT.models.Customer import Customer
from DO_AN_NHOM_KTLT.models.Customer_Room import Customer_Room
from DO_AN_NHOM_KTLT.models.Room import Room


class ExportTool:
    def export_room_to_excel(self,filename,rooms):
        workbook = xr.Workbook(filename)
        worksheet = workbook.add_worksheet()

        # Modify column width
        worksheet.set_column('A:A', 15)
        worksheet.set_column('B:B', 25)
        worksheet.set_column('C:C', 15)
        worksheet.set_column('D:D', 15)
        worksheet.set_column('E:E', 15)
        bold = workbook.add_format({'bold': True})
        #Add header
        worksheet.write('A1', 'ID phòng', bold)
        worksheet.write('B1', 'Loại phòng', bold)
        worksheet.write('C1', 'Giá/giờ', bold)
        worksheet.write('D1', 'Giá/ngày', bold)
        worksheet.write('E1', 'Trạng thái', bold)
        worksheet.write('F1', 'Tầng', bold)
        for i in range(len(rooms)):
            index = i + 2
            r = rooms[i]
            worksheet.write(f'A{index}', r.room_id)
            worksheet.write(f'B{index}', r.room_type)
            worksheet.write(f'C{index}', r.price_hourly)
            worksheet.write(f'D{index}', r.price_overnight)
            worksheet.write(f'E{index}', r.status)
            worksheet.write(f'F{index}', r.floor)
        workbook.close()
    def export_customer_to_excel(self,filename,customers,customer_room):
        workbook = xr.Workbook(filename)
        worksheet = workbook.add_worksheet()

        # Modify column width
        worksheet.set_column('A:A', 15)
        worksheet.set_column('B:B', 25)
        worksheet.set_column('C:C', 15)
        worksheet.set_column('D:D', 15)
        worksheet.set_column('E:E', 15)
        worksheet.set_column('F:F', 15)

        bold = workbook.add_format({'bold': True})
        date_format = workbook.add_format({'num_format': 'dd-mm-yyyy HH:MM:SS'})
        #Add header
        worksheet.write('A1', 'Mã KH', bold)
        worksheet.write('B1', 'Tên KH', bold)
        worksheet.write('C1', 'SĐT', bold)
        worksheet.write('D1', 'CCCD', bold)
        worksheet.write('E1', 'Mã phòng', bold)
        worksheet.write('F1', 'Ngày nhận phòng', bold)
        row_index = 1  # Bắt đầu từ dòng 2

        for cr in customer_room:
            # Tìm thông tin khách hàng tương ứng
            customer = next((c for c in customers if c.customer_id == cr.customer_id), None)

            if customer:
                worksheet.write(row_index, 0, customer.customer_id)
                worksheet.write(row_index, 1, customer.name)
                worksheet.write(row_index, 2, customer.phone)
                worksheet.write(row_index, 3, customer.cccd)

            worksheet.write(row_index, 4, cr.room_id)
            # Kiểm tra kiểu dữ liệu của ngày nhận phòng
            if isinstance(cr.date_checkin, datetime):
                date_obj = cr.date_checkin  # Đã là datetime thì giữ nguyên
            else:
                date_obj = datetime.strptime(cr.date_checkin, "%Y-%m-%d %H:%M:%S")  # Chuyển đổi từ str
            worksheet.write_datetime(row_index, 5, date_obj, date_format)
            row_index += 1  # Xuống dòng tiếp theo
        workbook.close()
    def import_customer_excel(self,filename):
        wb = load_workbook(filename)
        ws = wb[wb.sheetnames[0]]
        is_header=True
        customers=[]
        customer_rooms=[]
        for row in ws.values:
            if is_header==True:
                is_header=False
                continue
            customer_id, name, phone, cccd, room_id, date_checkin = row
            cus=Customer(customer_id, name, phone, cccd)
            customers.append(cus)
            if room_id and date_checkin:
                customer_rooms.append(Customer_Room(customer_id, room_id, date_checkin))
        wb.close()
        return customers, customer_rooms


    def import_room_excel(self,filename):
        wb = load_workbook(filename)
        ws = wb[wb.sheetnames[0]]
        is_header = True
        rooms = []
        for row in ws.values:
            if is_header == True:
                is_header = False
                continue
            room_id = row[0]
            room_type = row[1]
            price_hourly=float(row[2])
            price_overnight=float(row[3])
            status=row[4]
            floor=row[5]
            r=Room(room_id,room_type,price_hourly,price_overnight,status,floor)
            rooms.append(r)
        wb.close()
        return rooms